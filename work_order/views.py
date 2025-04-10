import json
import datetime
from datetime import timezone
from decimal import Decimal

#django
from django.urls import reverse
from django.db.models import Q,Sum,Min,Max 
from django.db import transaction, IntegrityError
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.models import User, Group
from django.shortcuts import get_object_or_404, render,redirect
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from django.forms import formset_factory, inlineformset_factory
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Count, Sum
from django.db.models import Exists, OuterRef

# rest framework
from api.v1.customers.serializers import CustomerSerializer
from rest_framework import status
#local
from .forms import *
from .models import *
from settings.forms import *
from settings.models import *
from main.decorators import role_required
from main.functions import generate_form_errors, get_auto_id,log_activity
from django.core.paginator import Paginator, PageNotAnInteger,EmptyPage
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill

import pandas as pd

class ColorAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        qs = Color.objects.all()

        if self.q:
            qs = qs.filter(name__icontains=self.q)

        return qs

    def create_object(self, text):
        # Allow new colors to be created
        return Color.objects.create(name=text)

def fetch_customer_details(request):
    mobile_no = request.GET.get("customer_mobile_no")
    
    if (instances:=Customer.objects.filter(mobile_number=mobile_no,is_deleted=False)).exists():
        instance = instances.first()
        seralizer = CustomerSerializer(instance)
        
        status_code = status.HTTP_200_OK
        response_data = {
            "status": "true",
            "status_code": status.HTTP_200_OK,
            "data": seralizer.data,
        }
    else:
        status_code = status.HTTP_404_NOT_FOUND
        response_data = {
            "status": "false",
            "status_code": status.HTTP_404_NOT_FOUND,
        }

    return HttpResponse(json.dumps(response_data),status=status_code, content_type="application/json")

@login_required
# @role_required(['superadmin'])
def work_order_info(request,pk):
    """
    WorkOrder List
    :param request:
    :return: WorkOrder List single view
    """
    
    model_images_instances = {}
    instance = WorkOrder.objects.get(pk=pk)
    order_item=WorkOrderItems.objects.filter(work_order=instance).first()
    items_instances = WorkOrderItems.objects.filter(work_order=instance)
    images_instances = WorkOrderImages.objects.filter(work_order=instance)
    work_section_details=WorkOrderStaffAssign.objects.filter(work_order=instance)

    if order_item:
        if not images_instances:
            model_images_instances=ModelNumberBasedProductImages.objects.filter(model__model_no=order_item.model_no)
            

    context = {
        'instance': instance,
        'items_instances':items_instances,
        'images_instances': images_instances,
        'model_images':model_images_instances,
        'page_name' : 'Work Order Info',
        'page_title' : 'Work Order Info',
        'section_details':work_section_details,
    }

    return render(request, 'admin_panel/pages/work_order/order/info.html', context)

@login_required
# @role_required(['superadmin'])
def work_order_list(request):
    """
    work_order
    :param request:
    :return: work_order list view
    """
    filter_data = {}
    query = request.GET.get("q")
    
    instances = WorkOrder.objects.filter(is_deleted=False).order_by("-date_added")
    
    if query:
        instances = instances.filter(
            Q(order_no__icontains=query) |
            Q(customer__name__icontains=query) 
        )
        title = "work_order list - %s" % query
        filter_data['q'] = query
    
    context = {
        'instances': instances,
        'page_name' : 'WorkOrders List',
        'page_title' : 'WorkOrders List',
        'filter_data' :filter_data,
    }

    return render(request, 'admin_panel/pages/work_order/order/list.html', context)


@login_required
# @role_required(['superadmin'])
def create_work_order(request):
    WorkOrderItemsFormFormset = formset_factory(WorkOrderItemsForm, extra=2)
    WorkOrderImagesFormFormset = formset_factory(WorkOrderImagesForm, extra=2)

    if request.method == 'POST':
        customer_form = CustomerForm(request.POST, files=request.FILES)
        work_order_form = WorkOrderForm(request.POST)
        work_order_items_formset = WorkOrderItemsFormFormset(request.POST, prefix='work_order_items_formset', form_kwargs={'empty_permitted': False})
        work_order_images_formset = WorkOrderImagesFormFormset(request.POST, files=request.FILES, prefix='work_order_images_formset', form_kwargs={'empty_permitted': False})
        
        print(Customer.objects.filter(mobile_number=request.POST.get("mobile_number")).exists())
        
        if customer_form.is_valid() and work_order_form.is_valid() and work_order_items_formset.is_valid() and work_order_images_formset.is_valid():
            try:
                with transaction.atomic():
                    # Check if a customer with the given mobile number already exists
                    if (customer_data:=Customer.objects.filter(mobile_number=request.POST.get("mobile_number"))).exists():
                        customer_data = customer_data.first()
                    else:
                        # Create a new user for the customer
                        user_data = User.objects.create_user(
                            username=customer_form.cleaned_data.get("mobile_number"),
                            password=f'{customer_form.cleaned_data.get("name")}@123',
                            is_active=True,
                        )

                        # Add the user to the customer group
                        group, created = Group.objects.get_or_create(name="customer")
                        user_data.groups.add(group)

                        # Create a new customer instance
                        customer_data = customer_form.save(commit=False)
                        customer_data.auto_id = get_auto_id(Customer)
                        customer_data.creator = request.user
                        customer_data.user = user_data
                        customer_data.save()

                    # Create the work order
                    work_order_data = work_order_form.save(commit=False)
                    work_order_data.auto_id = get_auto_id(WorkOrder)
                    work_order_data.creator = request.user
                    work_order_data.customer = customer_data
                    work_order_data.save()

                    # Save work order items
                    for form in work_order_items_formset:
                        work_order_item = form.save(commit=False)
                        work_order_item.auto_id = get_auto_id(WorkOrderItems)
                        work_order_item.creator = request.user
                        work_order_item.work_order = work_order_data
                        work_order_item.save()
    
                        if not ModelNumberBasedProducts.objects.filter(model_no=work_order_item.model_no).exists():
                            model_number_based_product=ModelNumberBasedProducts.objects.create(
                                auto_id=get_auto_id(ModelNumberBasedProducts),
                                creator=request.user,
                                model_no=work_order_item.model_no,
                                category=work_order_item.category,
                                sub_category=work_order_item.sub_category,
                                material=work_order_item.material,
                                sub_material=work_order_item.sub_material,
                                material_type=work_order_item.material_type,
                                
                            )
                            model_number_based_product.color.add(work_order_item.color)
                            model_number_based_product.size.add(work_order_item.size)
                            model_number_based_product.save()
                        else:
                            model_number_based_product = ModelNumberBasedProducts.objects.get(model_no=work_order_item.model_no)
                            model_number_based_product.color.add(work_order_item.color)
                            model_number_based_product.size.add(work_order_item.size)
                            model_number_based_product.save()

                    # Save work order images
                    for form in work_order_images_formset:
                        work_order_image = form.save(commit=False)
                        work_order_image.auto_id = get_auto_id(WorkOrderImages)
                        work_order_image.creator = request.user
                        work_order_image.work_order = work_order_data
                        work_order_image.save()
                    log_activity(
                        created_by=request.user,
                        description=f"Created work order-- '{work_order_data}'"
                        )

                    response_data = {
                        "status": "true",
                        "title": "Successfully Created",
                        "message": "WorkOrders created successfully.",
                        'redirect': 'true',
                        "redirect_url": reverse('work_order:work_order_list')
                    }

            except IntegrityError as e:
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
        else:
            message = generate_form_errors(customer_form, formset=False)
            message += generate_form_errors(work_order_form, formset=False)
            message += generate_form_errors(work_order_items_formset, formset=True)
            message += generate_form_errors(work_order_images_formset, formset=True)

            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
                "form_errors": work_order_form.errors.as_json(),
                "formset_errors": work_order_items_formset.errors,
            }

        return JsonResponse(response_data)

    else:
        customer_form = CustomerForm()
        work_order_form = WorkOrderForm(initial={'order_no': WorkOrder.generate_order_no()})
        work_order_items_formset = WorkOrderItemsFormFormset(prefix='work_order_items_formset')
        work_order_images_formset = WorkOrderImagesFormFormset(prefix='work_order_images_formset')

        context = {
            'customer_form': customer_form,
            'work_order_form': work_order_form,
            'work_order_items_formset': work_order_items_formset,
            'work_order_images_formset': work_order_images_formset,
            'page_name': 'Create WorkOrder',
            'page_title': 'Create WorkOrders',
            'url': reverse('work_order:create_work_order'),
            'work_order_page': True,
            'is_need_select2': True,
        }

        return render(request, 'admin_panel/pages/work_order/order/create.html', context)

    
@login_required
# @role_required(['superadmin'])
def edit_work_order(request, pk):
    """
    Edit operation of work_order
    :param request:
    :param pk:
    :return:
    """
    work_order_instance = get_object_or_404(WorkOrder, pk=pk)
    work_order_items = WorkOrderItems.objects.filter(work_order=work_order_instance, is_deleted=False)
    work_order_images = WorkOrderImages.objects.filter(work_order=work_order_instance, is_deleted=False)

    item_extra = 0 if work_order_items.exists() else 1
    image_extra = 0 if work_order_images.exists() else 1

    WorkOrderItemsFormFormset = inlineformset_factory(
        WorkOrder,
        WorkOrderItems,
        extra=item_extra,
        form=WorkOrderItemsForm,
    )
    
    WorkOrderImageFormFormset = inlineformset_factory(
        WorkOrder,
        WorkOrderImages,
        extra=image_extra,
        form=WorkOrderImagesForm,
    )

    message = ''

    if request.method == 'POST':
        customer_form = CustomerForm(request.POST, files=request.FILES, instance=work_order_instance.customer)
        work_order_form = WorkOrderForm(request.POST, files=request.FILES, instance=work_order_instance)
        work_order_items_formset = WorkOrderItemsFormFormset(
            request.POST,
            instance=work_order_instance,
            prefix='work_order_items_formset',
            form_kwargs={'empty_permitted': False}
        )
        work_order_images_formset = WorkOrderImageFormFormset(
            request.POST,
            instance=work_order_instance,
            prefix='work_order_images_formset',
            form_kwargs={'empty_permitted': False},
            files=request.FILES
        )

        if customer_form.is_valid() and work_order_form.is_valid() and work_order_items_formset.is_valid() and work_order_images_formset.is_valid():
            try:
                with transaction.atomic():
                    # Use cleaned_data now after form validation
                    mobile_number = customer_form.cleaned_data.get("mobile_number")

                    if Customer.objects.filter(mobile_number=mobile_number).exists():
                        customer_data = Customer.objects.get(mobile_number=mobile_number)
                    else:
                        user_data = User.objects.create_user(
                            username=mobile_number,
                            password=f'{customer_form.cleaned_data.get("name")}@123',
                            is_active=True,
                        )

                        group, created = Group.objects.get_or_create(name="customer")
                        user_data.groups.add(group)
                        
                        customer_data = customer_form.save(commit=False)
                        customer_data.auto_id = get_auto_id(Customer)
                        customer_data.creator = request.user
                        customer_data.user = user_data
                        customer_data.save()
                        
                    work_order_form_instance = work_order_form.save(commit=False)
                    work_order_form_instance.date_updated = datetime.now()
                    work_order_form_instance.updater = request.user
                    work_order_form_instance.save()
                    
                    for form in work_order_items_formset:
                        if form not in work_order_items_formset.deleted_forms:
                            work_order_item = form.save(commit=False)
                            if not work_order_item.auto_id:
                                work_order_item.work_order = work_order_form_instance
                                work_order_item.auto_id = get_auto_id(WorkOrderItems)
                                work_order_item.creator=request.user
                            work_order_item.updater = request.user
                            work_order_item.date_updated = datetime.now()
                            work_order_item.save()
                            
                            if not ModelNumberBasedProducts.objects.filter(model_no=work_order_item.model_no).exists():
                                product=ModelNumberBasedProducts.objects.create(
                                    auto_id = get_auto_id(ModelNumberBasedProducts),
                                    creator = request.user,
                                    model_no = work_order_item.model_no,
                                    category = work_order_item.category,
                                    sub_category = work_order_item.sub_category,
                                    material = work_order_item.material,
                                    sub_material = work_order_item.sub_material,
                                    material_type = work_order_item.material_type,
                                    #color = work_order_item.color,
                                )
                                product.color.add(form.cleaned_data['color'])
                                product.size.add(form.cleaned_data['size'])
                                product.save()
                

                    for form in work_order_items_formset.deleted_forms:
                        form.instance.delete()
                        
                    for form in work_order_images_formset:
                        if form not in work_order_images_formset.deleted_forms:
                            work_order_image = form.save(commit=False)
                            if not work_order_image.auto_id:
                                work_order_image.work_order = work_order_form_instance
                                work_order_image.auto_id = get_auto_id(WorkOrderImages)
                                work_order_image.creator=request.user
                            work_order_image.updater = request.user
                            work_order_image.date_updated = datetime.now()
                            work_order_image.save()
                            
                    for form in work_order_images_formset.deleted_forms:
                        form.instance.delete()

                    log_activity(
                        created_by=request.user,
                        description=f"Updated work order-- '{work_order_instance}'"
                        )

                    response_data = {
                        "status": "true",
                        "title": "Successfully Updated",
                        "message": "WorkOrders updated successfully.",
                        "redirect": "true",
                        "redirect_url": reverse('work_order:work_order_list'),
                    }

            except IntegrityError as e:
                # Handle database integrity error
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                # Handle other exceptions
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

        else:
            message = generate_form_errors(customer_form, formset=False)
            message += generate_form_errors(work_order_form, formset=False)
            message += generate_form_errors(work_order_items_formset, formset=True)
            message += generate_form_errors(work_order_images_formset, formset=True)

            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')

    else:
        customer_form = CustomerForm(instance=work_order_instance.customer)
        work_order_form = WorkOrderForm(instance=work_order_instance)
        work_order_items_formset = WorkOrderItemsFormFormset(
            queryset=work_order_items,
            prefix='work_order_items_formset',
            instance=work_order_instance
        )
        work_order_images_formset = WorkOrderImageFormFormset(
            queryset=work_order_images,
            prefix='work_order_images_formset',
            instance=work_order_instance
        )
        
        context = {
            'customer_form': customer_form,
            'work_order_form': work_order_form,
            'work_order_items_formset': work_order_items_formset,
            'work_order_images_formset': work_order_images_formset,
            'message': message,
            'page_name': 'edit work order',
            'is_purchase_pages': True,
            'is_purchase_page': True,
        }

        return render(request, 'admin_panel/pages/work_order/order/create.html', context)


@login_required
# @role_required(['superadmin'])
def delete_work_order(request,pk):
    """
    work_order deletion, it only mark as is deleted field to true
    :param request:
    :param pk:
    :return:
    """
    work_order_instance = get_object_or_404(WorkOrder, pk=pk)
    WorkOrderItems.objects.filter(work_order=work_order_instance,is_deleted=False).update(is_deleted=True)
    #WorkOrderImages.objects.filter(work_order=work_order_instance,is_deleted=False).update(is_deleted=True)
    
    work_order_instance.is_deleted=True
    work_order_instance.save()
    log_activity(
                created_by=request.user,
                description=f"Deleted work order-- '{work_order_instance}'"
            )
    
    response_data = {
        "status": "true",
        "title": "Successfully Deleted",
        "message": "WorkOrder Successfully Deleted.",
        "redirect": "true",
        "redirect_url": reverse('work_order:work_order_list'),
    }

    return redirect('work_order:work_order_list')
    
    #return HttpResponse(json.dumps(response_data), content_type='application/javascript')


@login_required
# @role_required(['superadmin'])
def delete_work_order_image(request,pk):
    """
    work_order deletion, it only mark as is deleted field to true
    :param request:
    :param pk:
    :return:
    """
    image = get_object_or_404(WorkOrderImages, id=pk)
    image.delete()
    log_activity(
                created_by=request.user,
                description=f"Deleted work order image-- '{image}'"
            )
    # if request.method == "POST":
    #     print('idd--------',pk)
        # image = get_object_or_404(WorkOrderImages, id=pk)
        # image.delete()
    #     return JsonResponse({"success": True})
    # return JsonResponse({"success": False})
    response_data = {
        "status": "true",
        "title": "Successfully Deleted",
        "message": "Work Order Image Successfully Deleted.",
        "redirect": "true",
        "redirect_url": reverse('work_order:model-display'),
    }

    
    return HttpResponse(json.dumps(response_data), content_type='application/javascript')

@login_required
# # @role_required(['superadmin'])
def assign_work_order(request):
    """
    Assign operation of work_order
    :param request:
    :param pk:
    :return:
    """
    if request.method == 'POST':
        form = WorkOrderStatusForm(request.POST)

        if form.is_valid() :
            try:
                with transaction.atomic():
                    # Update data
                    work_order = WorkOrder.objects.get(pk=request.POST.get('order_id'))
                    
                    data = form.save(commit=False)
                    data.auto_id = get_auto_id(WorkOrderStatus)
                    data.creator = request.user
                    data.work_order = work_order
                    data.from_section = work_order.status
                    data.save()
                    
                    work_order.status = data.to_section
                    work_order.save()
                    log_activity(
                        created_by=request.user,
                        description=f"Assigned work order --'{work_order}'"
                        )

                    response_data = {
                        "status": "true",
                        "title": "Successfully Updated",
                        "message": "Assign Work Orders updated successfully.",
                        "redirect": "true",
                        "redirect_url": reverse('work_order:work_order_list'),
                    }

            except IntegrityError as e:
                # Handle database integrity error
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                # Handle other exceptions
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

        else:
            message = generate_form_errors(form, formset=False)

            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')

@login_required
# @role_required(['superadmin'])
def profit_loss(request, pk):
    work_order = get_object_or_404(WorkOrder, pk=pk)
    
    customer = work_order.customer
    
    work_order_items = WorkOrderItems.objects.filter(work_order=work_order)
    
    # Calculate total estimated cost
    total_estimate = work_order_items.aggregate(Sum('estimate_rate'))['estimate_rate__sum'] or Decimal(0)
    
    # Fetch total wages from WorkOrderStaffAssign
    total_wages = WorkOrderStaffAssign.objects.filter(work_order=work_order).aggregate(Sum('wage'))['wage__sum'] or Decimal(0)
    # Add total_price calculation for each item
    for item in work_order_items:
        item.total_price = item.quantity * item.estimate_rate
    # Determine profit or loss
    profit_or_loss = total_estimate - total_wages
    status = "Profit" if profit_or_loss > 0 else "Loss"

    context = {
        'work_order': work_order,
        'customer': customer,
        'work_order_items': work_order_items,
        'total_estimate': total_estimate,
        'total_wages': total_wages,
        'profit_or_loss': profit_or_loss,
        'status': status,
    }
    return render(request, 'admin_panel/pages/work_order/order/profit_loss.html', context)


#----------------------------Wood Section---------------------------
def wood_work_orders_list(request):
    filter_data = {}
    query = request.GET.get("q")
    work_order_ids = WorkOrderStatus.objects.filter(to_section="012").values_list("work_order__pk")
    
    work_orders = WorkOrder.objects.filter(pk__in=work_order_ids,is_deleted=False)
    
    
    if query:
        work_orders = work_orders.filter(
            Q(order_no__icontains=query) |
            Q(customer__name__icontains=query) 
        )
        title = "work_order list - %s" % query
        filter_data['q'] = query
        
    context = {
        'page_name' : 'Wood Work Orders',
        'page_title': 'Wood Work Orders',
        'work_orders': work_orders,
        'filter_data' :filter_data,
    }
    
    return render(request, 'admin_panel/pages/wood/list.html', context) 

def assign_wood(request, pk):
    
    WoodWorksAssignFormSet = formset_factory(WoodWorksAssignForm, extra=1)
    work_order = get_object_or_404(WorkOrder, id=pk)

    if request.method == 'POST':
        formset = WoodWorksAssignFormSet(request.POST, request.FILES, prefix='formset', form_kwargs={'empty_permitted': False})
        message = ''
        if formset.is_valid():
            try:
                with transaction.atomic():
                    # Save each form in the formset
                    for form in formset:
                        wood_assign = form.save(commit=False)
                        wood_assign.work_order = work_order
                        wood_assign.auto_id = get_auto_id(WoodWorkAssign)
                        wood_assign.creator = request.user
                        wood_assign.work_order.status = "012"
                        wood_assign.save()
                    
                    work_order.is_assigned = True
                    work_order.save()
                    log_activity(
                        created_by=request.user,
                        description=f"Assigned wood for  '{work_order}'"
                        )

                    response_data = {
                        "status": "true",
                        "title": "Successfully Assigned",
                        "message": "Wood assigned successfully.",
                        "redirect": "true",
                        "redirect_url": reverse('work_order:wood_work_orders_list')
                    }

            except IntegrityError as e:
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
            except Exception as e:
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
        else:
            message = generate_form_errors(formset, formset=True)
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        formset = WoodWorksAssignFormSet(prefix='formset')
        context = {
            'formset': formset,
            'page_name': 'Wood Assign',
            'page_title': 'Wood Assign',
            'work_order': work_order,
            'url': reverse('work_order:assign_wood', args=[pk]),
            'is_need_select2': True,
        }
        return render(request, 'admin_panel/pages/wood/assign_wood.html', context)

def allocated_wood(request, pk):
    
    work_order = get_object_or_404(WorkOrder, id=pk)
    wood_assign = WoodWorkAssign.objects.filter(work_order=work_order)
    
    context = {
        'work_order': work_order,
        'wood_assign': wood_assign,
    }

    html = render_to_string('admin_panel/pages/wood/allocated_wood.html', context, request=request)
    return JsonResponse({'html': html})

#---------------------Carpentary Section----------------------------------
def carpentary_list(request):
    filter_data = {}
    query = request.GET.get("q")
    work_order_ids = WorkOrderStatus.objects.filter(to_section="015").values_list("work_order__pk")
    work_orders = WorkOrder.objects.filter(pk__in=work_order_ids,is_deleted=False)
    
    if query:
        work_orders = work_orders.filter(
            Q(order_no__icontains=query) |
            Q(customer__name__icontains=query) 
        )
        title = "work_order list - %s" % query
        filter_data['q'] = query
        
    context = {
        'page_name' : 'Carpentary',
        'page_title': 'Carpentary',
        'carpentary': work_orders,
        'filter_data' :filter_data,
    }
    
    return render(request, 'admin_panel/pages/wood/carpentary_list.html', context) 



def assign_carpentary(request, pk):
    work_order = get_object_or_404(WorkOrder, id=pk)
    
    CarpentaryAssignFormSet = inlineformset_factory(
        WorkOrder, Carpentary, 
        form=CarpentaryAssignForm, 
        extra=1, 
        can_delete=True
    )
    
    if request.method == 'POST':
        formset = CarpentaryAssignFormSet(request.POST, request.FILES, instance=work_order, prefix='formset')
        message = ''

        if formset.is_valid():
            try:
                with transaction.atomic():
                    instances = formset.save(commit=False)
                    for instance in instances:
                        instance.auto_id = get_auto_id(Carpentary)
                        instance.creator = request.user
                        instance.save()
                    work_order.status = "015"
                    work_order.is_assigned = True
                    work_order.save()
                    log_activity(
                        created_by=request.user,
                        description=f"Assigned carpentary for '{work_order}'"
                        )

                    response_data = {
                        "status": "true",
                        "title": "Successfully Assigned",
                        "message": "Carpentary assigned successfully.",
                        "redirect": "true",
                        "redirect_url": reverse('work_order:carpentary_list')
                    }
            except IntegrityError as e:
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
            except Exception as e:
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
        else:
            message = generate_form_errors(formset, formset=True)
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        formset = CarpentaryAssignFormSet(instance=work_order, prefix='formset')
        context = {
            'carpentary_formset': formset,
            'page_name': 'Carpentary Assign',
            'page_title': 'Carpentary Assign',
            'work_order': work_order,
            'url': reverse('work_order:assign_carpentary', args=[pk]),
            'is_need_select2': True,
        }
        
        return render(request, 'admin_panel/pages/wood/assign_carpentary.html', context)
    

def allocated_carpentary(request, pk):
    
    work_order = get_object_or_404(WorkOrder, id=pk)
    carpentary = Carpentary.objects.filter(work_order=work_order)
    
    context = {
        'work_order': work_order,
        'assign_carpentary': carpentary,
    }

    html = render_to_string('admin_panel/pages/wood/allocated_carpentary.html', context, request=request)
    return JsonResponse({'html': html})

#-----------------------Polish-------------------------------------------------------------------

def polish_list(request):
    filter_data = {}
    query = request.GET.get("q")
    work_order_ids = WorkOrderStatus.objects.filter(to_section="018").values_list("work_order__pk")
    work_orders = WorkOrder.objects.filter(pk__in=work_order_ids,is_deleted=False)
    if query:
        work_orders = work_orders.filter(
            Q(order_no__icontains=query) |
            Q(customer__name__icontains=query) 
        )
        title = "work_order list - %s" % query
        filter_data['q'] = query
    context = {
        'page_name' : 'Polish',
        'page_title': 'Polish',
        'polish': work_orders,
        'filter_data' :filter_data,
    }
    
    return render(request, 'admin_panel/pages/polish/polish_list.html', context) 

def assign_polish(request, pk):
    work_order = get_object_or_404(WorkOrder, id=pk)
    
    PolishAssignFormSet = inlineformset_factory(
        WorkOrder, Polish, 
        form=PolishAssignForm, 
        extra=1, 
        can_delete=True
    )
    
    if request.method == 'POST':
        formset = PolishAssignFormSet(request.POST, request.FILES, instance=work_order, prefix='formset')
        message = ''

        if formset.is_valid():
            try:
                with transaction.atomic():
                    instances = formset.save(commit=False)
                    for instance in instances:
                        instance.auto_id = get_auto_id(Polish)
                        instance.creator = request.user
                        instance.save()
                    work_order.status = "018"
                    work_order.is_assigned = True
                    work_order.save()
                    log_activity(
                        created_by=request.user,
                        description=f"Assigned polish for '{work_order}'"
                        )

                    response_data = {
                        "status": "true",
                        "title": "Successfully Assigned",
                        "message": "Polish assigned successfully.",
                        "redirect": "true",
                        "redirect_url": reverse('work_order:polish_list')
                    }
            except IntegrityError as e:
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
            except Exception as e:
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
        else:
            message = generate_form_errors(formset, formset=True)
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        formset = PolishAssignFormSet(instance=work_order, prefix='formset')
        context = {
            'polish_formset': formset,
            'page_name': 'Polish Assign',
            'page_title': 'Polish Assign',
            'work_order': work_order,
            'url': reverse('work_order:assign_polish', args=[pk]),
            'is_need_select2': True,
        }
        
        return render(request, 'admin_panel/pages/polish/assign_polish.html', context)
    

def allocated_polish(request, pk):
    
    work_order = get_object_or_404(WorkOrder, id=pk)
    polish = Polish.objects.filter(work_order=work_order)
    
    context = {
        'work_order': work_order,
        'assign_polish': polish,
    }

    html = render_to_string('admin_panel/pages/polish/allocated_polish.html', context, request=request)
    return JsonResponse({'html': html})


#--------------------------------Glass----------------------------------------------------


def glass_list(request):
    filter_data = {}
    query = request.GET.get("q")
    work_order_ids = WorkOrderStatus.objects.filter(to_section="020").values_list("work_order__pk")
    work_orders = WorkOrder.objects.filter(pk__in=work_order_ids,is_deleted=False)
    if query:
        work_orders = work_orders.filter(
            Q(order_no__icontains=query) |
            Q(customer__name__icontains=query) 
        )
        title = "work_order list - %s" % query
        filter_data['q'] = query
    context = {
        'page_name' : 'Glass',
        'page_title': 'Glass',
        'glass': work_orders,
        'filter_data' :filter_data,
    }
    
    return render(request, 'admin_panel/pages/glass/glass_list.html', context) 

def assign_glass(request, pk):
    work_order = get_object_or_404(WorkOrder, id=pk)
    
    GlassAssignFormSet = inlineformset_factory(
        WorkOrder, Glass, 
        form=GlassAssignForm, 
        extra=1, 
        can_delete=True
    )
    
    if request.method == 'POST':
        formset = GlassAssignFormSet(request.POST, request.FILES, instance=work_order, prefix='formset')
        message = ''

        if formset.is_valid():
            try:
                with transaction.atomic():
                    instances = formset.save(commit=False)
                    for instance in instances:
                        instance.auto_id = get_auto_id(Glass)
                        instance.creator = request.user
                        instance.save()
                    work_order.status = "020"
                    work_order.is_assigned = True
                    work_order.save()
                    log_activity(
                        created_by=request.user,
                        description=f"Assigned glass for  '{work_order}'"
                        )

                    response_data = {
                        "status": "true",
                        "title": "Successfully Assigned",
                        "message": "Glass assigned successfully.",
                        "redirect": "true",
                        "redirect_url": reverse('work_order:glass_list')
                    }
            except IntegrityError as e:
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
            except Exception as e:
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
        else:
            message = generate_form_errors(formset, formset=True)
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        formset = GlassAssignFormSet(instance=work_order, prefix='formset')
        context = {
            'glass_formset': formset,
            'page_name': 'Glass Assign',
            'page_title': 'Glass Assign',
            'work_order': work_order,
            'url': reverse('work_order:assign_glass', args=[pk]),
            'is_need_select2': True,
        }
        
        return render(request, 'admin_panel/pages/glass/assign_glass.html', context)
    

def allocated_glass(request, pk):
    
    work_order = get_object_or_404(WorkOrder, id=pk)
    glass = Glass.objects.filter(work_order=work_order)
    
    context = {
        'work_order': work_order,
        'assign_glass': glass,
    }

    html = render_to_string('admin_panel/pages/glass/allocated_glass.html', context, request=request)
    return JsonResponse({'html': html})


#----------------------------Packing--------------------------------------------------------------------
def packing_list(request):
    filter_data = {}
    query = request.GET.get("q")
    work_order_ids = WorkOrderStatus.objects.filter(to_section="022").values_list("work_order__pk")
    work_orders = WorkOrder.objects.filter(pk__in=work_order_ids,is_deleted=False)
    if query:
        work_orders = work_orders.filter(
            Q(order_no__icontains=query) |
            Q(customer__name__icontains=query) 
        )
        title = "work_order list - %s" % query
        filter_data['q'] = query
    context = {
        'page_name' : 'Packing',
        'page_title': 'Packing',
        'packing': work_orders,
        'filter_data' :filter_data,
    }
    
    return render(request, 'admin_panel/pages/packing/packing_list.html', context) 

def assign_packing(request, pk):
    work_order = get_object_or_404(WorkOrder, id=pk)
    
    PackingAssignFormSet = inlineformset_factory(
        WorkOrder, Packing, 
        form=PackingAssignForm, 
        extra=1, 
        can_delete=True
    )
    
    if request.method == 'POST':
        formset = PackingAssignFormSet(request.POST, request.FILES, instance=work_order, prefix='formset')
        message = ''

        if formset.is_valid():
            try:
                with transaction.atomic():
                    instances = formset.save(commit=False)
                    for instance in instances:
                        instance.auto_id = get_auto_id(Packing)
                        instance.creator = request.user
                        instance.save()
                    work_order.status = "022"
                    work_order.is_assigned = True
                    work_order.save()
                    log_activity(
                       created_by=request.user,
                       description=f"Assigned packing for '{work_order}'"
                    )

                    response_data = {
                        "status": "true",
                        "title": "Successfully Assigned",
                        "message": "Packing assigned successfully.",
                        "redirect": "true",
                        "redirect_url": reverse('work_order:packing_list')
                    }
            except IntegrityError as e:
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
            except Exception as e:
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
        else:
            message = generate_form_errors(formset, formset=True)
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        formset = PackingAssignFormSet(instance=work_order, prefix='formset')
        context = {
            'packing_formset': formset,
            'page_name': 'Packing Assign',
            'page_title': 'Packing Assign',
            'work_order': work_order,
            'url': reverse('work_order:assign_packing', args=[pk]),
            'is_need_select2': True,
        }
        
        return render(request, 'admin_panel/pages/packing/assign_packing.html', context)
    

def allocated_packing(request, pk):
    
    work_order = get_object_or_404(WorkOrder, id=pk)
    packing = Packing.objects.filter(work_order=work_order)
    
    context = {
        'work_order': work_order,
        'assign_packing': packing,
    }

    html = render_to_string('admin_panel/pages/packing/allocated_packing.html', context, request=request)
    return JsonResponse({'html': html})



#---------------------------------------- assigning staff in wood section----------------------------#
def wood_order_staff_assign(request,pk):

    work_order=get_object_or_404(WorkOrder,id=pk)

    if request.method=="GET":
        assigned_staffs=WorkOrderStaffAssign.objects.filter(work_order=work_order)
        staff=Staff.objects.filter(department__name="Wood Section")

        return render(request,'admin_panel/pages/work_order/order/staff_assign.html',{"instance":work_order,"staff":staff,"staffs":assigned_staffs})
    
    elif request.method=="POST":
        staff_id=request.POST.get("staff")
        time=request.POST.get("time")
        wage=request.POST.get("wage")
        staff=Staff.objects.get(id=staff_id)
        WorkOrderStaffAssign.objects.create(staff=staff,work_order=work_order,
                                            creator=request.user,auto_id=get_auto_id(WorkOrderStaffAssign),
                                            time_spent=time,wage=wage
                                            )
        log_activity(
                created_by=request.user,
                description=f"Assigned '{staff}' for'{work_order}'"
            )

        return redirect('work_order:wood_work_orders_list')
    
    
#------------- assigning staff in carpentary section---------------------#
def carpentary_order_staff_assign(request,pk):

    work_order=get_object_or_404(WorkOrder,id=pk)

    if request.method=="GET":
        staff=Staff.objects.filter(department__name="Carpentary")
        assigned_staffs=WorkOrderStaffAssign.objects.filter(work_order=work_order)

        return render(request,'admin_panel/pages/work_order/order/staff_assign.html',{"instance":work_order,"staff":staff,"staffs":assigned_staffs})
    
    elif request.method=="POST":
        staff_id=request.POST.get("staff")
        time=request.POST.get("time")
        wage=request.POST.get("wage")
        staff=Staff.objects.get(id=staff_id)
        WorkOrderStaffAssign.objects.create(staff=staff,work_order=work_order,
                                            creator=request.user,auto_id=get_auto_id(WorkOrderStaffAssign),
                                            time_spent=time,wage=wage
                                            )
        log_activity(
                created_by=request.user,
                description=f"Assigned '{staff}' for '{work_order}'"
            )

        return redirect('work_order:carpentary_list')
    

#-----------------------assiging staff for polish section--------------------------------#
def polish_order_staff_assign(request,pk):

    work_order=get_object_or_404(WorkOrder,id=pk)

    if request.method=="GET":
        assigned_staffs=WorkOrderStaffAssign.objects.filter(work_order=work_order)
        staff=Staff.objects.filter(department__name="Polish")

        return render(request,'admin_panel/pages/work_order/order/staff_assign.html',{"instance":work_order,"staff":staff,"staffs":assigned_staffs})
    
    elif request.method=="POST":
        staff_id=request.POST.get("staff")
        time=request.POST.get("time")
        wage=request.POST.get("wage")
        staff=Staff.objects.get(id=staff_id)
        
        WorkOrderStaffAssign.objects.create(staff=staff,work_order=work_order,
                                            creator=request.user,auto_id=get_auto_id(WorkOrderStaffAssign),
                                            time_spent=time,wage=wage
                                            )
        log_activity(
                created_by=request.user,
                description=f"Assigned '{staff}' for '{work_order}'"
            )


        return redirect('work_order:polish_list')    

#----------------------------------assigning staff for glass/upholstory section-----------#
def glass_order_staff_assign(request,pk):

    work_order=get_object_or_404(WorkOrder,id=pk)

    if request.method=="GET":
        staff=Staff.objects.filter(department__name="Glass/Upholstory")
        assigned_staffs=WorkOrderStaffAssign.objects.filter(work_order=work_order)

        return render(request,'admin_panel/pages/work_order/order/staff_assign.html',{"instance":work_order,"staff":staff,"staffs":assigned_staffs})
    
    elif request.method=="POST":
        staff_id=request.POST.get("staff")
        time=request.POST.get("time")
        wage=request.POST.get("wage")
        staff=Staff.objects.get(id=staff_id)
        
        WorkOrderStaffAssign.objects.create(staff=staff,work_order=work_order,
                                            creator=request.user,auto_id=get_auto_id(WorkOrderStaffAssign),
                                            time_spent=time,wage=wage
                                            )
        log_activity(
                created_by=request.user,
                description=f"Assigned '{staff}' for '{work_order}'"
            )

        return redirect('work_order:glass_list')       


#-------------------------assigning staff for packing section-------------#
def packing_order_staff_assign(request,pk):

    work_order=get_object_or_404(WorkOrder,id=pk)

    if request.method=="GET":
        staff=Staff.objects.filter(department__name="Packing")
        assigned_staffs=WorkOrderStaffAssign.objects.filter(work_order=work_order)

        return render(request,'admin_panel/pages/work_order/order/staff_assign.html',{"instance":work_order,"staff":staff,"staffs":assigned_staffs})
    
    elif request.method=="POST":
        staff_id=request.POST.get("staff")
        time=request.POST.get("time")
        wage=request.POST.get("wage")
        staff=Staff.objects.get(id=staff_id)
        
        WorkOrderStaffAssign.objects.create(staff=staff,work_order=work_order,
                                            creator=request.user,auto_id=get_auto_id(WorkOrderStaffAssign),
                                            time_spent=time,wage=wage
                                            )
        log_activity(
                created_by=request.user,
                description=f"Assigned '{staff}' for '{work_order}'"
            )
        

        return redirect('work_order:packing_list')
    

@login_required
# @role_required(['superadmin'])    
def create_color(request):
    
    if request.method == 'POST':
        form = ColorForm(request.POST)
        
        if form.is_valid():
            color = form.save(commit=False)
            color.creator = request.user 
            color.auto_id = get_auto_id(Color)
            color.save()
            log_activity(
                created_by=request.user,
                description=f"created '{color}'"
            )
        
            response_data = {
                "status": "true",
                "title": "Successfully Created",
                "message": "Color created successfully.",
                "redirect": "true",
                "redirect_url": reverse('work_order:color_list') 
            }
            return HttpResponse(json.dumps(response_data), content_type='application/javascript')
        else:
            message = generate_form_errors(form, formset=False)
        
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message
            }
            return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    else:
        form = ColorForm()

        context = {
            'form': form,
            'page_name': 'Create Color',
            'page_title': 'Create Color',
            'url': reverse('work_order:create_color')  # URL for form action.
        }
        return render(request, 'admin_panel/pages/work_order/create_color.html', context)


@login_required
# @role_required(['superadmin'])
def color_list(request):
    colors = Color.objects.all()

    context = {
        'colors': colors,
        'page_name': 'Color List',
        'page_title': 'List of Colors',
        'url': redirect('work_order:color_list'),  
    }

    return render(request, 'admin_panel/pages/work_order/color_list.html', context)

@login_required
# @role_required(['superadmin'])
def delete_color(request, pk):
    color = get_object_or_404(Color, pk=pk)
    
    color.delete()
    log_activity(
                created_by=request.user,
                description=f"Deleted '{color}'"
            )

    response_data = {
            "status": "true",
            "title": "Successfully Deleted",
            "message": "Color deleted successfully.",
            "redirect": "true",
            "redirect_url": reverse('work_order:color_list')  
        }
    return JsonResponse(response_data)
    

@login_required
# @role_required(['superadmin'])
def size_list(request):
    sizes = Size.objects.all()
    return render(request, 'admin_panel/pages/work_order/size_list.html', {'sizes': sizes})


@login_required
# @role_required(['superadmin'])
def size_create(request):
    if request.method == 'POST':
        form = SizeForm(request.POST)
        size=request.POST.get('size')
        if Size.objects.filter(Q(size__iexact=size)).exists():
            form.add_error('size', 'Size already exists.')
        else:
            if form.is_valid():
                form.save()
                log_activity(
                created_by=request.user,
                description=f"created '{size}'"
                )
                return redirect('work_order:size-list')
    else:
        form = SizeForm()
    return render(request, 'admin_panel/pages/work_order/size_create.html', {'form': form})


@login_required
# @role_required(['superadmin'])
def size_delete(request, pk):
    size= get_object_or_404(Size, pk=pk)
    
    size.delete()

    log_activity(
            created_by=request.user,
            description=f"Deleted '{size}'"
        )
    response_data = {
            "status": "true",
            "title": "Successfully Deleted",
            "message": "Size deleted successfully.",
            "redirect": "true",
            "redirect_url": reverse('work_order:size-list')  
        }
    return JsonResponse(response_data)


# @login_required
# # @role_required(['superadmin'])
# def modelnumberbasedproducts_list(request):
    
#     instances = ModelNumberBasedProducts.objects.all().order_by("-id")
    
#     context = {
#         'instances': instances,
#         'page_name': 'ModelNumberBasedProducts List',
#         'page_title': 'ModelNumberBasedProducts List',
#     }

#     return render(request, 'admin_panel/pages/work_order/model_list.html', context)
@login_required
# @role_required(['superadmin'])
def modelnumberbasedproducts_create(request):
    response_data = {}
    title='Model Create'
    
    WorkOrderImagesFormSet =formset_factory(ModelNumberBasedProductImagesForm,extra=2)
    
    if request.method == 'POST':
        
        form = ModelNumberBasedProductsForm(request.POST)
        work_order_image_formset = WorkOrderImagesFormSet(request.POST, files=request.FILES,prefix='work_order_image_formset',form_kwargs={'empty_permitted': False})
        model_no = request.POST.get('model_no')

        if ModelNumberBasedProducts.objects.filter(model_no=model_no).exists():
            # form.add_error('model_no', 'Model number already exists.')
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": "Model No : Model number already exists",
            }
        else:
            if form.is_valid() and work_order_image_formset.is_valid():
                try:
                    with transaction.atomic():
                        product = form.save(commit=False)
                        product.auto_id = get_auto_id(ModelNumberBasedProducts)
                        product.creator = request.user
                        product.save()
                        
                        product.color.set(form.cleaned_data['color'])
                        product.size.set(form.cleaned_data['size'])
                        product.save()
        
                        for form in work_order_image_formset:
                            image = form.save(commit=False)
                            image.model = product
                            image.auto_id = get_auto_id(ModelNumberBasedProductImages)
                            image.creator = request.user
                            image.save()
                        
                        log_activity(
                        created_by=request.user,
                        description=f"created '{product}'"
                        )
                        
                        response_data = {
                            "status": "true",
                            "title": "Successfully Created",
                            "message": "Modelnumber based Product created successfully.",
                            'redirect': 'true',
                            "redirect_url": reverse('work_order:model-display')
                            }
                except IntegrityError as e:
                    response_data = {
                        "status": "false",
                        "title": "Failed",
                        "message": str(e),
                    }

                except Exception as e:
                    response_data = {
                        "status": "false",
                        "title": "Failed",
                        "message": str(e),
                    }
            else:
                message = generate_form_errors(form, formset=False)
                message += generate_form_errors(work_order_image_formset, formset=True)
                
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": message,
                }
            
        return HttpResponse(json.dumps(response_data), content_type='application/javascript')

    else:
        form = ModelNumberBasedProductsForm()
        work_order_image_formset = WorkOrderImagesFormSet(prefix='work_order_image_formset')

    return render(request, 'admin_panel/pages/work_order/model_create.html', {'form': form, 'work_order_image_formset': work_order_image_formset,'title':title})


# @login_required
# # @role_required(['superadmin'])
# def modelnumberbasedproducts_update(request,pk):
#     product = get_object_or_404(ModelNumberBasedProducts,pk=pk)
#     title='Model Update'
#     WorkOrderImagesFormSet =formset_factory(WorkOrderImagesForm,extra=2)
    
#     if request.method == 'POST':
#         form = ModelNumberBasedProductsForm(request.POST, instance=product)
#         work_order_image_formset = WorkOrderImagesFormSet(request.POST, files=request.FILES,prefix='work_order_image_formset',form_kwargs={'empty_permitted': False})
#         model_no = request.POST.get('model_no')

#         if not ModelNumberBasedProducts.objects.filter(model_no=model_no).exists():
#             form.add_error('model_no', 'Model number does not exists.')
#         else:
#             if form.is_valid() and work_order_image_formset.is_valid():
#                 product = ModelNumberBasedProducts.objects.filter(model_no=form.cleaned_data['model_no']).update(
#                     auto_id=get_auto_id(ModelNumberBasedProducts),
#                     updater=request.user,
#                     date_updated = datetime.datetime.now(),
#                     model_no=form.cleaned_data['model_no'],
#                     category=form.cleaned_data['category'],
#                     sub_category=form.cleaned_data['sub_category'],
#                     material=form.cleaned_data['material'],
#                     sub_material=form.cleaned_data['sub_material'],
#                     material_type=form.cleaned_data['material_type']
#                 )
#                 product = ModelNumberBasedProducts.objects.get(pk=pk)
#                 product.color.set(form.cleaned_data['color'])
#                 product.size.set(form.cleaned_data['size'])
#                 product.save()

#                 for form in work_order_image_formset:
#                     if form.cleaned_data:
#                         image = form.save(commit=False)
#                         image.model = product
#                         image.auto_id = get_auto_id(ModelNumberBasedProductImages)
#                         image.creator = request.user
#                         image.save()
#                 log_activity(
#                 created_by=request.user,
#                 description=f"Updated model number based product for modelno- '{product}'"
#                 )

#                 response_data = {
#                         "status": "true",
#                         "title": "Successfully Updated",
#                         "message": "Modelnumber based Product updated successfully.",
#                         'redirect': 'true',
#                         "redirect_url": reverse('work_order:model-display')
#                         }
            
#             return HttpResponse(json.dumps(response_data), content_type='application/javascript')
#     else:
#         form = ModelNumberBasedProductsForm(instance=product)
#         work_order_image_formset = WorkOrderImagesFormSet()

#     return render(request, 'admin_panel/pages/work_order/model_create.html', {'form': form, 'work_order_image_formset': work_order_image_formset,'title':title})
@login_required
def modelnumberbasedproducts_update(request, pk):
    product = get_object_or_404(ModelNumberBasedProducts, pk=pk)
    title = 'Model Update'
    WorkOrderImagesFormSet = formset_factory(ModelNumberBasedProductImagesForm, extra=2)

    if request.method == 'POST':
        form = ModelNumberBasedProductsForm(request.POST, instance=product)
        work_order_image_formset = WorkOrderImagesFormSet(
            request.POST, files=request.FILES,
            prefix='work_order_image_formset', form_kwargs={'empty_permitted': False}
        )
        model_no = request.POST.get('model_no')

        response_data = {
            "status": "false",
            "title": "Update Failed",
            "message": "There was an error updating the product."
        }

        if not ModelNumberBasedProducts.objects.filter(model_no=model_no).exists():
            form.add_error('model_no', 'Model number does not exist.')

        if form.is_valid() or work_order_image_formset.is_valid():
            try:
                # Update product instance properly
                product.model_no = form.cleaned_data['model_no']
                product.category = form.cleaned_data['category']
                product.sub_category = form.cleaned_data['sub_category']
                product.material = form.cleaned_data['material']
                product.sub_material = form.cleaned_data['sub_material']
                product.material_type = form.cleaned_data['material_type']
                product.updater = request.user
                product.date_updated = datetime.now()
                product.save()

                # Many-to-Many relations should be updated AFTER saving the model
                product.color.set(form.cleaned_data['color'])
                product.size.set(form.cleaned_data['size'])

                # Save WorkOrderImages if provided
                for form in work_order_image_formset:
                    if form.cleaned_data:
                        image = form.save(commit=False)
                        image.model = product
                        image.auto_id = get_auto_id(ModelNumberBasedProductImages)
                        image.creator = request.user
                        image.save()

                log_activity(
                    created_by=request.user,
                    description=f"Updated model number based product for model_no '{product.model_no}'"
                )

                response_data = {
                    "status": "true",
                    "title": "Successfully Updated",
                    "message": "Model number based product updated successfully.",
                    "redirect": "true",
                    "redirect_url": reverse('work_order:model-display')
                }

            except Exception as e:
                response_data["message"] = f"Update failed: {str(e)}"

        else:
            response_data["message"] = f"Form errors: {form.errors.as_json()}"

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')

    else:
        form = ModelNumberBasedProductsForm(instance=product)
        work_order_image_formset = WorkOrderImagesFormSet()

    return render(request, 'admin_panel/pages/work_order/model_create.html', {
        'form': form,
        'work_order_image_formset': work_order_image_formset,
        'title': title
    })


@login_required
# @role_required(['superadmin'])
def modelnumberbasedproducts_delete(request, pk):
    model = get_object_or_404(ModelNumberBasedProducts, pk=pk)
    model.delete()
    log_activity(
                created_by=request.user,
                description=f"Deleted model number based product for model number '{model}'"
            )
    

    return redirect('work_order:model-display')


@login_required
# @role_required(['superadmin'])
def modelnumberbasedproducts_info(request,pk):
    model=get_object_or_404(ModelNumberBasedProducts,pk=pk)
    images=ModelNumberBasedProductImages.objects.filter(model=model)

    context = {
        'model':model,
        'images_instances': images,
        'page_name' : 'Model Info',
        'page_title' : 'Model Info',
    }

    return render(request, 'admin_panel/pages/work_order/model_info.html', context)



def get_model_details(request):
    model_no = request.GET.get('model_no')
    try:
        model = ModelNumberBasedProducts.objects.get(model_no=model_no)
        images=ModelNumberBasedProductImages.objects.filter(model__model_no=model_no)
        image_urls = [image.image.url for image in images]
        data = {
            'category': model.category.id,
            'sub_category': model.sub_category.id if model.sub_category else '',
            'material': model.material.id,
            'sub_material': model.sub_material.id if model.sub_material else '',
            'material_type': model.material_type.id if model.material_type else '',
            'color': list(model.color.values_list('id', flat=True)),
            'size': list(model.size.values_list('id', flat=True)),
            'images':image_urls,
        }
        
    except ModelNumberBasedProducts.DoesNotExist:
        data = {}
        
    return JsonResponse(data)


@login_required
# @role_required(['superadmin'])
def modelnumberbasedproducts_card_list(request):
    category_id = request.GET.get('category')
    sub_category_id = request.GET.get('sub_category')
    model_no=request.GET.get('q')

    work_orders = {}

    if category_id and sub_category_id:
        work_order_images=ModelNumberBasedProductImages.objects.filter(model__category=category_id).filter(model__sub_category=sub_category_id)
        print(work_order_images)
    elif category_id:
        work_order_images=ModelNumberBasedProductImages.objects.filter(model__category=category_id)
        print(work_order_images)
    elif sub_category_id:
        work_order_images=ModelNumberBasedProductImages.objects.filter(model__sub_category=sub_category_id)
        print(work_order_images)
    elif model_no:
        work_order_images=ModelNumberBasedProductImages.objects.filter(model__model_no=model_no)
    else:
        work_order_images = ModelNumberBasedProductImages.objects.all().order_by("-id")

    paginator = Paginator(work_order_images,6) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)


    for image in page_obj:
        if image.model not in work_orders:
            work_orders[image.model] = image

    categories = ProductCategory.objects.all()
    sub_categories = ProductSubCategory.objects.all

    for itmes in work_orders:
        print(itmes)
    
    context = {
        'work_orders': work_orders,
        'categories': categories,
        'sub_categories': sub_categories,
        'page_name':'Modelnumberbasedproducts',
        'page_obj': page_obj 
    }

    return render(request, 'admin_panel/pages/work_order/model_display.html',context)


def get_subcategories(request):
    category_id = request.GET.get('category')
    sub_categories = ProductSubCategory.objects.filter(product_category=category_id)

    return render(request, 'admin_panel/pages/work_order/subcategory_options.html', {'sub_categories': sub_categories})


def get_subcategory(request):
    category_id = request.GET.get('category')
    subcategories = ProductSubCategory.objects.filter(product_category=category_id).values('id', 'name')
    return JsonResponse(list(subcategories), safe=False)

def get_sub_materials(request):
    material_id = request.GET.get('material')
    sub_materials = MaterialsType.objects.filter(material_id=material_id).values('id', 'name')
    return JsonResponse(list(sub_materials), safe=False)

def get_material_types(request):
    sub_material_id = request.GET.get('sub_material')
    material_types = MaterialTypeCategory.objects.filter(material_type=sub_material_id).values('id', 'name')
    return JsonResponse(list(material_types), safe=False)



@login_required
# @role_required(['superadmin'])
def delete_model_image(request,pk):
    if request.method == "POST":
        image = get_object_or_404(ModelNumberBasedProductImages, id=pk)
        image.delete()
        log_activity(
                created_by=request.user,
                description=f"Deleted model image for model number '{image.model}'"
            )
        return JsonResponse({"success": True})
    return JsonResponse({"success": False})


@login_required
def delete_orders(request):
    if request.method == 'POST':
        order_ids = request.POST.getlist('order_ids')
        if order_ids:
            WorkOrder.objects.filter(id__in=order_ids).delete()
            log_activity(
                created_by=request.user,
                description=f"Deleted work orders for ids '{order_ids}'"
            )
            return redirect('work_order:work_order_list')


@login_required
# @role_required(['superadmin'])
def staff_work_order_report(request):
    """
    staff's work order report
    :param request:
    :return: staff's work order report list view
    """
    filter_data = {}
    query = request.GET.get("q")
    
    instances = Staff.objects.filter(is_deleted=False).order_by("-date_added")
    
    if query:
        instances = instances.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(phone__icontains=query) |
            Q(employee_id__icontains=query) 
        )
        title = "work order list - %s" % query
        filter_data['q'] = query
    
    context = {
        'instances': instances,
        'page_name' : 'Work Order List',
        'page_title' : 'Work Order List',
        'filter_data' :filter_data,
    }

    return render(request, 'admin_panel/pages/reports/work_report.html', context)


@login_required
# @role_required(['superadmin'])
def delayed_work_order_report(request):
    """
    delayed work order report
    :param request:
    :return: delayed work order report list view
    """
    filter_data = {}
    query = request.GET.get("q")
    
    instances = WorkOrder.objects.filter(delivery_date__lt=datetime.today().date(),is_deleted=False).exclude(status="030").order_by("-date_added")
    
    if query:
        instances = instances.filter(
            Q(order_no__icontains=query) |
            Q(customer__name__icontains=query) 
        )
        title = "delayed work order list - %s" % query
        filter_data['q'] = query
    
    context = {
        'instances': instances,
        'page_name' : 'Delayed Work Order List',
        'page_title' : 'Delayed Work Order List',
        'filter_data' :filter_data,
    }

    return render(request, 'admin_panel/pages/reports/delayed_order_list.html', context)


@login_required
def print_delayed_work_order_report(request):
    instances = WorkOrder.objects.filter(delivery_date__lt=datetime.today().date(), is_deleted=False).exclude(status="030").order_by("-date_added")

    query = request.GET.get("q")
    if query:
        instances = instances.filter(
            Q(order_no__icontains=query) |
            Q(customer__name__icontains=query)
        )

    context = {
        'instances': instances,
        'page_title': 'Print - Delayed Work Order Report',
    }
    return render(request, 'admin_panel/pages/reports/delayed_order_print.html', context)


@login_required
def export_delayed_work_orders_excel(request):
    """ Export delayed work orders as an Excel file """
    instances = WorkOrder.objects.filter(delivery_date__lt=datetime.today().date(), is_deleted=False).exclude(status="030")

    data = []
    for instance in instances:
        categories = ", ".join([item.category.name for item in instance.workorderitems_set.all()])
        data.append([instance.order_no, instance.customer.name, instance.customer.mobile_number, instance.number_of_items(), categories, instance.delivery_date, instance.delayed_days()])

    df = pd.DataFrame(data, columns=["WO No", "Client Name", "Mobile", "No of Items", "Item Category", "Planned Delivery", "Delayed Days"])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="delayed_work_orders.xlsx"'
    
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Delayed Orders', index=False)
    
    return response

@login_required
# @role_required(['superadmin'])
def work_summary(request):
    """
    delayed work order report
    :param request:
    :return: delayed work order report list view
    """
    filter_data = {}
    query = request.GET.get("q")
    
    instances = WorkOrder.objects.filter(delivery_date__lt=datetime.today().date(),is_deleted=False).exclude(status="030").order_by("-date_added")
    
    if query:
        instances = instances.filter(
            Q(order_no__icontains=query) |
            Q(customer__name__icontains=query) |
            Q(customer__mobile_number__icontains=query)
        )
        title = "work order summary list - %s" % query
        filter_data['q'] = query
        
    # Calculate total sum of total_estimate
    total_estimate_sum = instances.aggregate(Sum('total_estimate'))['total_estimate__sum'] or 0

    context = {
        'instances': instances,
        'page_name' : 'Work Order Summary List',
        'page_title' : 'Work Order Summary List',
        'filter_data' :filter_data,
        'total_estimate_sum' :total_estimate_sum,
    }

    return render(request, 'admin_panel/pages/reports/work_summary.html', context)

@login_required
# @role_required(['superadmin'])
def print_work_summary_report(request):
    instances = WorkOrder.objects.filter(delivery_date__lt=datetime.today().date(), is_deleted=False).exclude(status="030").order_by("-date_added")

    query = request.GET.get("q")
    if query:
        instances = instances.filter(
            Q(order_no__icontains=query) |
            Q(customer__name__icontains=query) |
            Q(customer__mobile_number__icontains=query)
        )
    # Calculate total sum of total_estimate
    total_estimate_sum = instances.aggregate(Sum('total_estimate'))['total_estimate__sum'] or 0
    context = {
        'instances': instances,
        'page_title': 'Print -  Work Order Report',
        'total_estimate_sum' :total_estimate_sum,
        
    }
    return render(request, 'admin_panel/pages/reports/work_summary_print.html', context)


@login_required
# @role_required(['superadmin'])
def export_work_orders_summary_excel(request):
    """ Export delayed work orders as an Excel file with footer and styling """

    query = request.GET.get("q")
    instances = WorkOrder.objects.filter(delivery_date__lt=datetime.today().date(), is_deleted=False).exclude(status="030")

    if query:
        instances = instances.filter(
            Q(order_no__icontains=query) |
            Q(customer__name__icontains=query) |
            Q(customer__mobile_number__icontains=query)
        )

    data = []
    for index, instance in enumerate(instances, start=1):
        data.append([
            index,  # Serial number
            instance.order_no,
            instance.customer.name,
            instance.customer.mobile_number,
            instance.number_of_items(),
            instance.delivery_date,
            instance.get_status_display(),
            instance.total_estimate
        ])

    # Convert to DataFrame
    df = pd.DataFrame(data, columns=[
        "#", "WO No", "Client Name", "Mobile Number", "No of Items",
        "Delivery Date", "Current Stage", "Order Value"
    ])

    # Calculate total sum of 'total_estimate' and append footer
    total_estimate_sum = instances.aggregate(Sum('total_estimate'))['total_estimate__sum'] or 0
    footer = [""] * 7 + [total_estimate_sum]
    df.loc[len(df)] = footer

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="work_order_summary.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Work Orders', index=False)

        # Apply styling
        workbook = writer.book
        sheet = writer.sheets['Work Orders']

        # Set header style
        header_font = Font(bold=True)
        for col in range(1, len(df.columns) + 1):
            sheet.cell(row=1, column=col).font = header_font

        # Highlight delayed orders
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        for row_idx, instance in enumerate(instances, start=2):  # Start from row 2 (after headers)
            if instance.delayed_days() > 0:
                for col_idx in range(1, len(df.columns) + 1):
                    sheet.cell(row=row_idx, column=col_idx).fill = red_fill

        # Apply footer style (bold text)
        footer_row = len(df)  # Footer row is the last row
        for col_idx in range(1, len(df.columns) + 1):
            sheet.cell(row=footer_row + 1, column=col_idx).font = Font(bold=True)

    return response


@login_required
# @role_required(['superadmin'])
def accessories_utilized(request):
    today = datetime.today().date()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    query = request.GET.get("q")

    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            start_date_obj = end_date_obj = today
    else:
        start_date_obj = end_date_obj = today

    filters = Q(is_deleted=False) & Q(work_order__is_deleted=False) & Q(date_added__date__range=(start_date_obj, end_date_obj))

    if query:
        filters &= (
            Q(material__name__icontains=query) |
            Q(material_type__name__icontains=query)
        )

    instances = WoodWorkAssign.objects.select_related(
        'work_order', 'material', 'material_type'
    ).filter(filters).order_by('-date_added')

    # Directly sum if fields are numeric
    total_quantity = instances.aggregate(total=Sum('quantity'))['total'] or 0
    total_cost = instances.aggregate(total=Sum('rate'))['total'] or 0

    context = {
        'instances': instances,
        'total_quantity': total_quantity,
        'total_cost': total_cost,
        'page_name': 'Accessories Utilized Report',
        'page_title': 'Accessories Utilized Report',
        'filter_data': {'q': query} if query else {},
        'start_date': start_date_obj.strftime('%Y-%m-%d'),
        'end_date': end_date_obj.strftime('%Y-%m-%d'),
    }

    return render(request, 'admin_panel/pages/reports/accessories_utilized.html', context)
@login_required
# @role_required(['superadmin'])
def print_accessories_utilized(request):
    today = datetime.today().date()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    query = request.GET.get("q")

    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            start_date_obj = end_date_obj = today
    else:
        start_date_obj = end_date_obj = today

    filters = Q(is_deleted=False) & Q(work_order__is_deleted=False) & Q(date_added__date__range=(start_date_obj, end_date_obj))

    if query:
        filters &= (
            Q(material__name__icontains=query) |
            Q(material_type__name__icontains=query)
        )

    instances = WoodWorkAssign.objects.select_related(
        'work_order', 'material', 'material_type'
    ).filter(filters).order_by('-date_added')

    # Directly sum if fields are numeric
    total_quantity = instances.aggregate(total=Sum('quantity'))['total'] or 0
    total_cost = instances.aggregate(total=Sum('rate'))['total'] or 0

    context = {
        'instances': instances,
        'total_quantity': total_quantity,
        'total_cost': total_cost,
        'page_name': 'Accessories Utilized Report',
        'page_title': 'Accessories Utilized Report',
        'filter_data': {'q': query} if query else {},
        'start_date': start_date_obj.strftime('%Y-%m-%d'),
        'end_date': end_date_obj.strftime('%Y-%m-%d'),
    }

    return render(request, 'admin_panel/pages/reports/accessories_utilized_print.html', context)


@login_required
# @role_required(['superadmin'])
def export_accessories_utilized(request):
    today = datetime.today().date()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    query = request.GET.get("q")

    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            start_date_obj = end_date_obj = today
    else:
        start_date_obj = end_date_obj = today

    filters = Q(is_deleted=False) & Q(work_order__is_deleted=False) & Q(date_added__date__range=(start_date_obj, end_date_obj))

    if query:
        filters &= (
            Q(material__name__icontains=query) |
            Q(material_type__name__icontains=query)
        )

    instances = WoodWorkAssign.objects.select_related(
        'work_order', 'material', 'material_type'
    ).filter(filters).order_by('-date_added')

    # Totals
    total_quantity = instances.aggregate(total=Sum('quantity'))['total'] or 0
    total_cost = instances.aggregate(total=Sum('rate'))['total'] or 0

    # Build data for DataFrame
    data = []
    for idx, instance in enumerate(instances, start=1):
        data.append({
            '#': idx,
            'Date Added': instance.date_added.strftime('%d-%m-%Y') if instance.date_added else '',
            'Item Name': instance.material.name if instance.material else '',
            'Quantity Used': float(instance.quantity),
            'Cost': float(instance.rate),
            'Section Name': instance.work_order.get_status_display() if instance.work_order else '',
            'Delivery Date': instance.work_order.delivery_date.strftime('%d/%m/%Y') if instance.work_order and instance.work_order.delivery_date else ''
        })

    df = pd.DataFrame(data)

    # Prepare Excel response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="accessories_utilized.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Accessories Utilized', index=False)

        # Styling and Totals
        sheet = writer.sheets['Accessories Utilized']
        header_font = Font(bold=True)
        for col in range(1, len(df.columns) + 1):
            sheet.cell(row=1, column=col).font = header_font

        # Determine column indexes (1-based)
        item_col = df.columns.get_loc('Item Name') + 1
        qty_col = df.columns.get_loc('Quantity Used') + 1
        cost_col = df.columns.get_loc('Cost') + 1

        # Totals row (after data)
        total_row = len(df) + 2
        sheet.cell(row=total_row, column=item_col).value = 'Total:'
        sheet.cell(row=total_row, column=qty_col).value = total_quantity
        sheet.cell(row=total_row, column=cost_col).value = total_cost

        # Bold total row
        sheet.cell(row=total_row, column=item_col).font = Font(bold=True)
        sheet.cell(row=total_row, column=qty_col).font = Font(bold=True)
        sheet.cell(row=total_row, column=cost_col).font = Font(bold=True)

    return response

@login_required
# @role_required(['superadmin'])
def work_report(request):
    today = datetime.today().date()

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    query = request.GET.get("q")

    # Use today's date if not provided
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            start_date_obj = end_date_obj = today
    else:
        start_date_obj = end_date_obj = today

    filters = Q(work_order__is_deleted=False) & Q(date_added__date__range=(start_date_obj, end_date_obj))

    if query:
        filters &= (
            Q(staff__first_name__icontains=query) |
            Q(staff__last_name__icontains=query) |
            Q(staff__department__name__icontains=query)
        )

    instances = (
        WorkOrderStaffAssign.objects
        .select_related('work_order', 'staff', 'staff__department', 'staff__designation')
        .filter(filters)
        .values(
            'staff__first_name',
            'staff__last_name',
            'staff__department__name',
            'work_order__remark',
            'date_added',
        )
        .annotate(
            total_hours=Sum('time_spent'),
            total_wage=Sum('wage'),
            project_count=Count('work_order', distinct=True)
        )
        .order_by('-date_added')
    )

    total_hours = instances.aggregate(total=Sum('total_hours'))['total'] or 0
    total_project_count = instances.aggregate(total=Sum('project_count'))['total'] or 0
    total_wage = instances.aggregate(total=Sum('total_wage'))['total'] or 0

    context = {
        'instances': instances,
        'page_name': 'Work Report',
        'page_title': 'Work Report',
        'filter_data': {'q': query} if query else {},
        'start_date': start_date_obj.strftime('%Y-%m-%d'),
        'end_date': end_date_obj.strftime('%Y-%m-%d'),
        'total_hours': total_hours,
        'total_project_count': total_project_count,
        'total_wage': total_wage,
    }

    return render(request, 'admin_panel/pages/reports/work_report.html', context)

@login_required
# @role_required(['superadmin'])
def print_work_report(request):
    today = datetime.today().date()

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    query = request.GET.get("q")

    # Use today's date if not provided
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            start_date_obj = end_date_obj = today
    else:
        start_date_obj = end_date_obj = today

    filters = Q(work_order__is_deleted=False) & Q(date_added__date__range=(start_date_obj, end_date_obj))

    if query:
        filters &= (
            Q(staff__first_name__icontains=query) |
            Q(staff__last_name__icontains=query) |
            Q(staff__department__name__icontains=query)
        )

    instances = (
        WorkOrderStaffAssign.objects
        .select_related('work_order', 'staff', 'staff__department', 'staff__designation')
        .filter(filters)
        .values(
            'staff__first_name',
            'staff__last_name',
            'staff__department__name',
            'work_order__remark',
            'date_added',
        )
        .annotate(
            total_hours=Sum('time_spent'),
            total_wage=Sum('wage'),
            project_count=Count('work_order', distinct=True)
        )
        .order_by('-date_added')
    )

    total_hours = instances.aggregate(total=Sum('total_hours'))['total'] or 0
    total_project_count = instances.aggregate(total=Sum('project_count'))['total'] or 0
    total_wage = instances.aggregate(total=Sum('total_wage'))['total'] or 0

    context = {
        'instances': instances,
        'page_name': 'Work Report',
        'page_title': 'Work Report',
        'filter_data': {'q': query} if query else {},
        'start_date': start_date_obj.strftime('%Y-%m-%d'),
        'end_date': end_date_obj.strftime('%Y-%m-%d'),
        'total_hours': total_hours,
        'total_project_count': total_project_count,
        'total_wage': total_wage,
    }
    return render(request, 'admin_panel/pages/reports/work_report_print.html', context)

@login_required
# @role_required(['superadmin'])
def export_work_report_excel(request):
    today = datetime.today().date()

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    query = request.GET.get("q")

    # Parse dates
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            start_date_obj = end_date_obj = today
    else:
        start_date_obj = end_date_obj = today

    filters = Q(work_order__is_deleted=False) & Q(date_added__date__range=(start_date_obj, end_date_obj))

    if query:
        filters &= (
            Q(staff__first_name__icontains=query) |
            Q(staff__last_name__icontains=query) |
            Q(staff__department__name__icontains=query)
        )

    instances = (
        WorkOrderStaffAssign.objects
        .select_related('work_order', 'staff', 'staff__department', 'staff__designation')
        .filter(filters)
        .values(
            'staff__first_name',
            'staff__last_name',
            'staff__department__name',
            'work_order__remark',
            'date_added',
        )
        .annotate(
            total_hours=Sum('time_spent'),
            total_wage=Sum('wage'),
            project_count=Count('work_order', distinct=True)
        )
        .order_by('-date_added')
    )

    # Convert queryset to DataFrame
    df = pd.DataFrame(instances)
    df.rename(columns={
        'date_added': 'Date',
        'staff__first_name': 'First Name',
        'staff__last_name': 'Last Name',
        'staff__department__name': 'Section',
        'total_hours': 'Total hrs Engaged',
        'project_count': 'No of Projects Involved',
        'total_wage': 'Cost',
        'work_order__remark': 'Remark',
    }, inplace=True)

    # Format date
    if not df.empty:
        df['Date'] = pd.to_datetime(df['Date']).dt.strftime('%d/%m/%Y')

        # Add full name column if needed
        df['Staff Name'] = df['First Name'] + ' ' + df['Last Name']
        df.drop(columns=['First Name', 'Last Name'], inplace=True)

        # Rearrange columns
        df = df[['Date', 'Staff Name', 'Section', 'Total hrs Engaged', 'No of Projects Involved', 'Cost', 'Remark']]

        # Add total row
        total_row = pd.DataFrame({
            'Date': [''],
            'Staff Name': [''],
            'Section': ['Total'],
            'Total hrs Engaged': [df['Total hrs Engaged'].sum()],
            'No of Projects Involved': [df['No of Projects Involved'].sum()],
            'Cost': [df['Cost'].sum()],
            'Remark': ['']
        })
        df = pd.concat([df, total_row], ignore_index=True)

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="work_report.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Work Report', index=False)

        # Style header and total row
        sheet = writer.sheets['Work Report']
        header_font = Font(bold=True)

        # Bold header row
        for cell in sheet[1]:
            cell.font = header_font

        # Bold total row
        total_row_index = len(df.index)
        for cell in sheet[total_row_index]:
            cell.font = header_font

    return response

@login_required
# @role_required(['superadmin'])
def work_order_used_accessories_report(request):
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    selected_project = request.GET.get('project_name')
    search_query = request.GET.get('q', '').strip()

    today = datetime.today().date()
    yesterday = today - timedelta(days=1)

    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else yesterday
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today

    work_orders = WorkOrder.objects.filter(
        Q(id__in=WoodWorkAssign.objects.values('work_order')) |
        Q(id__in=Carpentary.objects.values('work_order')) |
        Q(id__in=Polish.objects.values('work_order')) |
        Q(id__in=Glass.objects.values('work_order')) |
        Q(id__in=Packing.objects.values('work_order')),
        is_deleted=False
    ).distinct().order_by('-date_added')

    if selected_project:
        work_orders = work_orders.filter(order_no=selected_project)

    results = []
    total_quantity = 0
    total_cost = 0
    total_rate = 0  # This will hold the sum of all rates

    for order in work_orders:
        combined_items = []

        for model in [WoodWorkAssign, Carpentary, Polish, Glass, Packing]:
            entries = model.objects.filter(
                work_order=order,
                is_deleted=False,
                date_added__date__range=(start_date, end_date)
            ).order_by('-date_added')

            for entry in entries:
                combined_items.append({
                    'date_added': entry.date_added,
                    'work_order': entry.work_order,
                    'material': entry.material,
                    'quantity': entry.quantity,
                    'rate': entry.rate,
                    'total': float(entry.quantity) * float(entry.rate),
                    'section': model.__name__,
                })

        if combined_items:
            latest_status = order.workorderstatus_set.order_by('-date_added').first()
            latest_section = latest_status.get_to_section_display() if latest_status else order.get_status_display()

            for item in combined_items:
                item['section'] = latest_section

            results.extend(combined_items)

    #  Search filter
    if search_query:
        results = [
            item for item in results if
            search_query.lower() in item['work_order'].order_no.lower() or
            search_query.lower() in item['material'].name.lower()
        ]

    # Totals after filtering
    for item in results:
        try:
            quantity = float(item['quantity'])
        except (ValueError, TypeError):
            quantity = 0
        try:
            rate = float(item['rate'])
        except (ValueError, TypeError):
            rate = 0

        total_quantity += quantity
        total_cost += item['total']
        total_rate += rate

    results.sort(key=lambda x: x['date_added'], reverse=True)

    #  Project dropdown list
    project_names = WorkOrder.objects.filter(
        Q(id__in=WoodWorkAssign.objects.values('work_order')) |
        Q(id__in=Carpentary.objects.values('work_order')) |
        Q(id__in=Polish.objects.values('work_order')) |
        Q(id__in=Glass.objects.values('work_order')) |
        Q(id__in=Packing.objects.values('work_order')),
        is_deleted=False
    ).values_list('order_no', flat=True).distinct()

    context = {
        'instances': results,
        'total_quantity': round(total_quantity, 2),
        'total_rate': round(total_rate, 2),
        'total_cost': round(total_cost, 2),
        'start_date': start_date,
        'end_date': end_date,
        'project_names': project_names,
        'selected_project': selected_project,
        'search_query': search_query,
    }

    return render(request, 'admin_panel/pages/reports/work_order_used_accessories_report.html', context)

@login_required
# @role_required(['superadmin'])
def print_work_order_used_accessories_report(request):
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    selected_project = request.GET.get('project_name')
    search_query = request.GET.get('q', '').strip()

    today = datetime.today().date()
    yesterday = today - timedelta(days=1)

    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else yesterday
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today

    work_orders = WorkOrder.objects.filter(
        Q(id__in=WoodWorkAssign.objects.values('work_order')) |
        Q(id__in=Carpentary.objects.values('work_order')) |
        Q(id__in=Polish.objects.values('work_order')) |
        Q(id__in=Glass.objects.values('work_order')) |
        Q(id__in=Packing.objects.values('work_order')),
        is_deleted=False
    ).distinct().order_by('-date_added')

    if selected_project:
        work_orders = work_orders.filter(order_no=selected_project)

    results = []
    total_quantity = 0
    total_cost = 0
    total_rate = 0  # This will hold the sum of all rates

    for order in work_orders:
        combined_items = []

        for model in [WoodWorkAssign, Carpentary, Polish, Glass, Packing]:
            entries = model.objects.filter(
                work_order=order,
                is_deleted=False,
                date_added__date__range=(start_date, end_date)
            ).order_by('-date_added')

            for entry in entries:
                combined_items.append({
                    'date_added': entry.date_added,
                    'work_order': entry.work_order,
                    'material': entry.material,
                    'quantity': entry.quantity,
                    'rate': entry.rate,
                    'total': float(entry.quantity) * float(entry.rate),
                    'section': model.__name__,
                })

        if combined_items:
            latest_status = order.workorderstatus_set.order_by('-date_added').first()
            latest_section = latest_status.get_to_section_display() if latest_status else order.get_status_display()

            for item in combined_items:
                item['section'] = latest_section

            results.extend(combined_items)

    #  Search filter
    if search_query:
        results = [
            item for item in results if
            search_query.lower() in item['work_order'].order_no.lower() or
            search_query.lower() in item['material'].name.lower()
        ]

    # Totals after filtering
    for item in results:
        try:
            quantity = float(item['quantity'])
        except (ValueError, TypeError):
            quantity = 0
        try:
            rate = float(item['rate'])
        except (ValueError, TypeError):
            rate = 0

        total_quantity += quantity
        total_cost += item['total']
        total_rate += rate

    results.sort(key=lambda x: x['date_added'], reverse=True)

    #  Project dropdown list
    project_names = WorkOrder.objects.filter(
        Q(id__in=WoodWorkAssign.objects.values('work_order')) |
        Q(id__in=Carpentary.objects.values('work_order')) |
        Q(id__in=Polish.objects.values('work_order')) |
        Q(id__in=Glass.objects.values('work_order')) |
        Q(id__in=Packing.objects.values('work_order')),
        is_deleted=False
    ).values_list('order_no', flat=True).distinct()

    context = {
        'instances': results,
        'total_quantity': round(total_quantity, 2),
        'total_rate': round(total_rate, 2),
        'total_cost': round(total_cost, 2),
        'start_date': start_date,
        'end_date': end_date,
        'project_names': project_names,
        'selected_project': selected_project,
        'search_query': search_query,
    }

    return render(request, 'admin_panel/pages/reports/wo_used_accessories_report.html', context)

@login_required
# @role_required(['superadmin'])
def export_work_order_used_accessories_report(request):
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    selected_project = request.GET.get('project_name')
    search_query = request.GET.get('q', '').strip()

    today = datetime.today().date()
    yesterday = today - timedelta(days=1)

    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else yesterday
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today

    work_orders = WorkOrder.objects.filter(
        Q(id__in=WoodWorkAssign.objects.values('work_order')) |
        Q(id__in=Carpentary.objects.values('work_order')) |
        Q(id__in=Polish.objects.values('work_order')) |
        Q(id__in=Glass.objects.values('work_order')) |
        Q(id__in=Packing.objects.values('work_order')),
        is_deleted=False
    ).distinct().order_by('-date_added')

    if selected_project:
        work_orders = work_orders.filter(order_no=selected_project)

    results = []
    total_quantity = 0
    total_cost = 0
    total_rate = 0

    for order in work_orders:
        combined_items = []
        for model in [WoodWorkAssign, Carpentary, Polish, Glass, Packing]:
            entries = model.objects.filter(
                work_order=order,
                is_deleted=False,
                date_added__date__range=(start_date, end_date)
            ).order_by('-date_added')
            for entry in entries:
                combined_items.append({
                    'date_added': entry.date_added,
                    'work_order': entry.work_order,
                    'material': entry.material,
                    'quantity': entry.quantity,
                    'rate': entry.rate,
                    'total': float(entry.quantity) * float(entry.rate),
                    'section': model.__name__,
                })

        if combined_items:
            latest_status = order.workorderstatus_set.order_by('-date_added').first()
            latest_section = latest_status.get_to_section_display() if latest_status else order.get_status_display()

            for item in combined_items:
                item['section'] = latest_section
            results.extend(combined_items)

    if search_query:
        results = [
            item for item in results if
            search_query.lower() in item['work_order'].order_no.lower() or
            search_query.lower() in item['material'].name.lower()
        ]

    data = []
    for i, instance in enumerate(results, start=1):
        row = {
            '#': i,
            'Order Added Date': instance['date_added'].strftime('%d-%m-%Y'),
            'Project Name': instance['work_order'].order_no,
            'Accessories Used': instance['material'].name,
            'Quantity': instance['quantity'],
            'Rate': instance['rate'],
            'Total': instance['total'],
            'Section Name': instance['section'],
        }
        total_quantity += float(instance['quantity'])
        total_rate += float(instance['rate'])        
        total_cost += float(instance['total'])
        data.append(row)

    df = pd.DataFrame(data)

    total_row = {
        '#': '',
        'Order Added Date': '',
        'Project Name': 'Total',
        'Accessories Used': '',
        'Quantity': round(total_quantity, 2),
        'Rate': round(total_rate, 2),
        'Total': round(total_cost, 2),
        'Section Name': ''
    }
    df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"work_order_used_accessories_report_{timezone.now().date()}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Used Accessories Report', index=False)
        sheet = writer.sheets['Used Accessories Report']

        bold_font = Font(bold=True)

        # Bold header row
        for cell in sheet[1]:
            cell.font = bold_font

        # Bold total row (last row in sheet)
        total_row_idx = df.shape[0] + 1  # +1 for header row
        for cell in sheet[total_row_idx]:
            cell.font = bold_font

    return response

def production_cost_wo_list(request):
    """
    WorkOrder list view filtered by date and search
    """
    filter_data = {}
    query = request.GET.get("q")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    # Default: Yesterday & Today
    if not start_date and not end_date:
        today = timezone.now().date()
        yesterday = today - timedelta(days=1)
        start_date = yesterday.strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

    # Base queryset
    instances = WorkOrder.objects.filter(is_deleted=False)

    # Date filter
    if start_date and end_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
            instances = instances.filter(date_added__date__range=(start, end))
        except ValueError:
            pass  # Invalid date format, ignore

    # Search filter
    if query:
        instances = instances.filter(
            Q(order_no__icontains=query) |
            Q(customer__name__icontains=query)
        )
        filter_data['q'] = query

    instances = instances.order_by("-date_added")
    
    # Totals
    total_items = sum([instance.number_of_items() for instance in instances])
    total_estimate = sum([instance.total_estimate for instance in instances])
    total_actual_cost = sum([instance.get_actual_cost() for instance in instances])

    context = {
        'instances': instances,
        'page_name': 'Production Cost Report List',
        'page_title': 'Production Cost Report List',
        'filter_data': filter_data,
        'start_date': start_date,
        'end_date': end_date,
        'total_items': total_items,
        'total_estimate': total_estimate,
        'total_actual_cost': total_actual_cost,
    }
    return render(request, 'admin_panel/pages/reports/production_cost_wo_list.html', context)

def production_cost_wo_print(request):
    """
    WorkOrder list view filtered by date and search
    """
    filter_data = {}
    query = request.GET.get("q")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    # Default: Yesterday & Today
    if not start_date and not end_date:
        today = timezone.now().date()
        yesterday = today - timedelta(days=1)
        start_date = yesterday.strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

    # Base queryset
    instances = WorkOrder.objects.filter(is_deleted=False)

    # Date filter
    if start_date and end_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
            instances = instances.filter(date_added__date__range=(start, end))
        except ValueError:
            pass  # Invalid date format, ignore

    # Search filter
    if query:
        instances = instances.filter(
            Q(order_no__icontains=query) |
            Q(customer__name__icontains=query)
        )
        filter_data['q'] = query

    instances = instances.order_by("-date_added")
    
    # Totals
    total_items = sum([instance.number_of_items() for instance in instances])
    total_estimate = sum([instance.total_estimate for instance in instances])
    total_actual_cost = sum([instance.get_actual_cost() for instance in instances])

    context = {
        'instances': instances,
        'page_name': 'Print-Production Cost Report',
        'page_title': 'Print-Production Cost Report',
        'filter_data': filter_data,
        'start_date': start_date,
        'end_date': end_date,
        'total_items': total_items,
        'total_estimate': total_estimate,
        'total_actual_cost': total_actual_cost,
    }
    return render(request, 'admin_panel/pages/reports/production_cost_wo_print.html', context)

def production_cost_wo_export(request):
    """
    Export Production Cost Report in Excel format
    """
    query = request.GET.get("q")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    # Default: Yesterday & Today
    if not start_date and not end_date:
        today = timezone.now().date()
        yesterday = today - timedelta(days=1)
        start_date = yesterday.strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

    # Base queryset
    instances = WorkOrder.objects.filter(is_deleted=False)

    # Date filter
    if start_date and end_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
            instances = instances.filter(date_added__date__range=(start, end))
        except ValueError:
            pass  # Invalid date format, ignore

    # Search filter
    if query:
        instances = instances.filter(
            Q(order_no__icontains=query) |
            Q(customer__name__icontains=query)
        )

    instances = instances.order_by("-date_added")

    # Prepare data for DataFrame
    data = []
    for i, instance in enumerate(instances, 1):
        data.append({
            "#": i,
            "Date Added": instance.date_added.strftime("%d-%m-%Y"),
            "Order No": instance.order_no,
            "Customer Name": instance.customer.name,
            "No Of Items": instance.number_of_items(),
            "Estimated Rate": float(instance.total_estimate),
            "Status": instance.get_status_display(),
            "Actual Cost": float(instance.get_actual_cost()),
            "Profit/Loss": instance.get_profit_or_loss(),
        })

    # Create DataFrame
    df = pd.DataFrame(data)

    # Add Total row
    total_row = {
        "#": "",
        "Date Added": "",
        "Order No": "Total",
        "Customer Name": "",
        "No Of Items": df["No Of Items"].sum(),
        "Estimated Rate": df["Estimated Rate"].sum(),
        "Status": "",
        "Actual Cost": df["Actual Cost"].sum(),
        "Profit/Loss": "",
    }
    df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"production_cost_report_{timezone.now().date()}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Export to Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Production Cost Report', index=False)
        sheet = writer.sheets['Production Cost Report']

        bold_font = Font(bold=True)

        # Bold header row
        for cell in sheet[1]:
            cell.font = bold_font

        # Bold total row
        total_row_idx = df.shape[0] + 1  # +1 for header row
        for cell in sheet[total_row_idx]:
            cell.font = bold_font

    return response